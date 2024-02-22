from django.shortcuts import render, redirect
from django.http import HttpResponse
from Techno.forms import *
from django.core.exceptions import ValidationError
from TechnoCoder.models import *
from django.contrib.auth import login, authenticate
from datetime import timedelta
from django.utils import timezone
from django.contrib.auth.signals import user_logged_in
from django.dispatch import receiver
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook, load_workbook  

# Create your views here.
def login_page(request):
    if(request.method=='POST'):
        username_or_email  = request.POST.get('username_or_email')
        password = request.POST.get('password')
        
        if '@' in username_or_email:
            user_obj = User.objects.filter(email=username_or_email)
        else:
            user_obj = User.objects.filter(username=username_or_email)
        
        if not user_obj.exists():
            error_message = "Username or Password is incorrect or doesn't exists"
            return render(request, 'login.html', {'error_message': error_message})
        
        username = username_or_email
        if '@' in username_or_email:
            username= User.objects.get(email=username_or_email).username
        user_obj = authenticate(request,username=username, password=password)

        if user_obj:
            login(request, user_obj)
            print(user_obj)
            url='/home/{0}'.format(request.user.username)
            return redirect(url)
             
        error_message = "Wrong Password"
        return render(request, 'login.html', {'error_message': error_message})
    return render(request, 'login.html')

def register(request):
    if(request.method=='POST'):
        form = Register(request.POST)
        if form.is_valid():
            form.save()  
            return HttpResponse("Thank you, You are registered now")
    else:
        form= Register()
    return render(request, 'register.html',{'form':form})

@login_required
def home(request,username):
    certificates = {'CompTIA A+', 'CompTIA Network+', 'Cisco Certified Network Associate (CCNA)', 'Cisco Certified Network Professional (CCNP)',
                'Certified Information Systems Security Professional (CISSP)', 'Certified ScrumMaster (CSM)',
                'Certified Data Professional (CDP)','Red Hat Certified Engineer (RHCE)','AWS Certified Developer – Associate',
                'Microsoft Certified: Azure Developer Associate','Google Associate Cloud Engineer','Microsoft Certified: Azure AI Engineer Associate',
                'IBM Data Science Professional Certificate'}
    if request.method =='GET':
        user=User.objects.get(username=username)
        userProfile = UserProfile.objects.filter(user=user)
        return render(request,'index.html',{'certificates':userProfile})
            
    return render(request,'index.html',{'certificates':certificates})

def profile(request):
    profile = UserProfile.objects.get(user=request.user)
    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('profile')
    else:
        form = UserForm(instance=profile)
    return render(request, 'profile.html', {'form': form})

def password_reset(request):
    return render(request,'forgot-password.html',{'template':'password_reset.html'})

@receiver(user_logged_in)
def log_user_login(sender, user, request, **kwargs):
    LoginActivity.objects.create(user=user)

def calculate_login_streak(user):
    today = timezone.now().date()
    streak = 0
    current_date = today
    max_streak = 0

    while LoginActivity.objects.filter(user=user, login_date=current_date).exists():
        streak += 1
        current_date -= timedelta(days=1)

    return streak

def importCSV(request):
    wb = load_workbook('C:/Users/Lenovo/Downloads/certificates questions.xlsx')
    ws = wb.active
    ws = wb['Cisco Certified Network Associate']
    for row in ws.iter_rows(min_row=2, values_only=True):
        id, question,choice1,choice2,choice3,choice4,correctChoice= row
        course=Course.objects.filter(title__contains='Cisco Certified Network Associate (CCNA)')
        if course.exists():
            course=Course.objects.get(title__contains='Cisco Certified Network Associate (CCNA)')
        Questions.objects.create(course=course,question=question,choice1=choice1,choice2=choice2,choice3=choice3,choice4=choice4,correctChoice=correctChoice)

    return HttpResponse("CSV imported Succesfully")
